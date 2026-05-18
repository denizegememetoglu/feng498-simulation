import json
import os
import re

import openpyxl

from src.config import DATA_FILE


_VALID_RACKS = frozenset({"A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "U"})
# Accept optional 'BR' prefix; özet uses both 'BRA-02-02' and 'A-02-02' interchangeably.
_BIN_RE = re.compile(r"^(?:BR)?([A-Z])-(\d{1,2})-(\d{1,2})$")
# Kardex automated storage — not in rack layout, tracked separately.
_KDX_RE = re.compile(r"^KDX\d*$")

# Last-load telemetry from load_storage_bins, surfaced via preprocess stats.
LAST_LOAD_META: dict[str, int] = {}


def _resolve_data_path():
    """Resolve the path to the Excel data file.

    Prefer the exact `DATA_FILE` from config when it exists (unicode
    normalisation can leave the filesystem name slightly different from the
    config string — try both raw and NFC). Falling back to the first .xlsx
    in the data/ directory is dangerous once a second Excel file (BOM,
    secondary export) is added, so we only do that as a last resort and
    raise if multiple candidates exist.
    """
    import unicodedata
    candidates = [DATA_FILE, unicodedata.normalize("NFC", DATA_FILE),
                  unicodedata.normalize("NFD", DATA_FILE)]
    for c in candidates:
        if os.path.exists(c):
            return c
    data_dir = os.path.dirname(DATA_FILE) or "data"
    xlsx_files = [f for f in os.listdir(data_dir) if f.endswith(".xlsx")]
    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx file found in {data_dir}/")
    if len(xlsx_files) > 1:
        raise FileNotFoundError(
            f"Multiple .xlsx files in {data_dir}/ ({xlsx_files}); "
            f"set DATA_FILE in src/config.py to disambiguate."
        )
    return os.path.join(data_dir, xlsx_files[0])


def _open_workbook(filepath=None):
    path = filepath or _resolve_data_path()
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def load_abc_analizi(filepath=None):
    """Load ABC Analizi sheet: 6011 rows of material classification + consumption."""
    wb = _open_workbook(filepath)
    ws = wb["ABC Analizi"]
    materials = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        mat_id, se_class, consumption, _team_abc_raw, pareto, abc, _se_abc_raw, _match_raw = row[:8]
        if not mat_id or not consumption or consumption <= 0:
            continue
        se_class = str(se_class).strip() if se_class else "CR"
        if len(se_class) < 2 or se_class[0] not in "ABC" or se_class[1] not in "FMRD":
            continue
        consumption = float(consumption)
        pareto = float(pareto) if pareto else 0.0
        if abc and str(abc).strip() in ("A", "B", "C"):
            team_abc = str(abc).strip()
        elif pareto <= 0.80:
            team_abc = "A"
        elif pareto <= 0.95:
            team_abc = "B"
        else:
            team_abc = "C"

        se_abc = se_class[0]
        se_fmr = se_class[1]

        materials.append({
            "material_id": str(mat_id).strip(),
            "se_class": se_class,
            "se_abc": se_abc,
            "se_fmr": se_fmr,
            "consumption": consumption,
            "pareto": pareto,
            "team_abc": team_abc,
            "abc_match": 1 if team_abc == se_abc else 0,
        })
    wb.close()
    materials.sort(key=lambda m: m["consumption"], reverse=True)
    return materials


def load_consumption(filepath=None):
    """Load zppq11 sheet: Material | Material number | Plnt | material-plant | TOTAL."""
    wb = _open_workbook(filepath)
    ws = wb["zppq11"]
    result: dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 5:
            continue
        mat_id = row[0]
        total = row[4]
        if not mat_id or total is None:
            continue
        try:
            t = float(total)
        except (TypeError, ValueError):
            continue
        if t > 0:
            result[str(mat_id).strip()] = t
    wb.close()
    return result


def load_sap_master(filepath=None):
    """Load zppq16_copy sheet: SAP material master data."""
    wb = _open_workbook(filepath)
    ws = wb["zppq16_copy"]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        mat_id, plant, abc_ind, purch_grp, mrp_ctrl, desc = row[:6]
        if not mat_id:
            continue
        result[str(mat_id).strip()] = {
            "plant": str(plant) if plant else "",
            "abc_indicator": str(abc_ind) if abc_ind else "",
            "purchasing_group": str(purch_grp) if purch_grp else "",
            "mrp_controller": str(mrp_ctrl) if mrp_ctrl else "",
            "description": str(desc) if desc else "",
        }
    wb.close()
    return result


def load_storage_bins(filepath=None):
    """Load özet sheet: material -> list of SAP storage bin codes.

    Returns `dict[str, list[str]]` so multi-bin materials (forward + reserve,
    or multiple slots for high-volume items) keep ALL their bins instead of
    silently collapsing to the last-seen one. Earlier versions kept only
    one bin per material and exposed the discarded duplicate count via
    LAST_LOAD_META — that's still done, but the data itself is no longer lossy.
    """
    wb = _open_workbook(filepath)
    ws = wb["özet"]
    result: dict[str, list[str]] = {}
    duplicates = 0
    conflicts = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Material | Plant | MRP Controller | MRP Tanımı | ABC Ind | Storage Bin
        if len(row) < 6:
            continue
        mat_id = row[0]
        bin_code = row[5]
        if not mat_id or not bin_code:
            continue
        mid = str(mat_id).strip()
        bc = str(bin_code).strip()
        bins = result.setdefault(mid, [])
        if bc in bins:
            duplicates += 1  # exact same bin repeated
        else:
            if bins:  # already had a different bin
                conflicts += 1
            bins.append(bc)
    wb.close()
    LAST_LOAD_META["bin_duplicates"] = duplicates
    LAST_LOAD_META["bin_conflicts"] = conflicts
    LAST_LOAD_META["multi_bin_materials"] = sum(1 for v in result.values() if len(v) > 1)
    return result


def is_kardex_bin(bin_code: str) -> bool:
    """Kardex automated storage (KDX, KDX2, KDX3, KDX4). Not in rack layout."""
    if not bin_code:
        return False
    return bool(_KDX_RE.match(str(bin_code).strip().upper()))


def decode_storage_bin(bin_code: str) -> tuple[str, int, int] | None:
    """Parse 'BRA-02-02' or 'A-02-02' into ('A', 2, 2). Returns None for malformed
    codes or codes with rack letters outside the documented 11 racks {A..I, J, U}.

    The convention was confirmed May 6 against the rack PDFs: optional 'BR'
    prefix, then the rack letter, then a 1-2 digit bay code, then a 1-2 digit
    position within the bay. Kardex codes (KDX*) are not racks and return None
    here — use is_kardex_bin to detect them separately.
    """
    if not bin_code:
        return None
    m = _BIN_RE.match(str(bin_code).strip().upper())
    if not m:
        return None
    rack = m.group(1)
    if rack not in _VALID_RACKS:
        return None
    bay = int(m.group(2))
    pos = int(m.group(3))
    return rack, bay, pos


def load_mrpc(filepath=None) -> dict[str, str]:
    """Load mrpc sheet: MRP-Controller code -> production line name (Açıklama)."""
    wb = _open_workbook(filepath)
    ws = wb["mrpc"]
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Plant | MRP-C | plant-mrpc | Açıklama
        if len(row) < 4:
            continue
        mrpc = row[1]
        line = row[3]
        if not mrpc or not line:
            continue
        result[str(mrpc).strip()] = str(line).strip()
    wb.close()
    return result


def preprocess(filepath=None, layout_path: str = None, write_stats: bool = True) -> dict:
    """Run all loaders, decode storage bins, join material -> line, and emit
    audit stats. Returns a single dict consumed by main.py.
    """
    materials = load_abc_analizi(filepath)
    consumption = load_consumption(filepath)
    sap_master = load_sap_master(filepath)
    storage_bins = load_storage_bins(filepath)
    mrp_to_line = load_mrpc(filepath)

    # Resolve layout for cross-validation (which (rack, bay) actually exist).
    from src.warehouse import Warehouse
    wh = Warehouse(layout_path)
    valid_rack_bays: set[tuple[str, int]] = set()
    for pid, p in wh.positions.items():
        valid_rack_bays.add((p.rack_id, p.bay_code))

    # Decoded bins is now a list per material (multi-bin support). A material
    # may have a mix of rack bins and Kardex codes — we keep the rack tuples
    # for placement and track kardex membership separately so the simulation
    # can route those picks to the Kardex resource.
    decoded_bins: dict[str, list[tuple[str, int, int]]] = {}
    kardex_materials: set[str] = set()
    bins_unknown_rack = 0
    bins_unmapped_position = 0
    bins_malformed = 0
    bins_kardex = 0

    for mat_id, bin_codes in storage_bins.items():
        for bin_code in bin_codes:
            if is_kardex_bin(bin_code):
                kardex_materials.add(mat_id)
                bins_kardex += 1
                continue
            decoded = decode_storage_bin(bin_code)
            if decoded is None:
                bins_malformed += 1
                continue
            rack, bay, pos = decoded
            if (rack, bay) not in valid_rack_bays:
                bins_unmapped_position += 1
                continue
            decoded_bins.setdefault(mat_id, []).append(decoded)

    # Join material -> line via sap_master.mrp_controller -> mrp_to_line
    material_to_line: dict[str, str] = {}
    for mat_id, master in sap_master.items():
        mrp = master.get("mrp_controller", "").strip()
        if not mrp:
            continue
        line = mrp_to_line.get(mrp)
        if line:
            material_to_line[mat_id] = line

    stats = {
        "materials_total": len(materials),
        "materials_with_bin": sum(1 for m in materials if m["material_id"] in storage_bins),
        "materials_with_decoded_bin": sum(1 for m in materials if m["material_id"] in decoded_bins),
        "materials_in_kardex": sum(1 for m in materials if m["material_id"] in kardex_materials),
        "bins_total": len(storage_bins),
        "bins_decoded": len(decoded_bins),
        "bins_kardex": bins_kardex,
        "bins_malformed": bins_malformed,
        "bins_unknown_rack": bins_unknown_rack,
        "bins_unmapped_position": bins_unmapped_position,
        "bin_duplicates": LAST_LOAD_META.get("bin_duplicates", 0),
        "bin_conflicts": LAST_LOAD_META.get("bin_conflicts", 0),
        "multi_bin_materials": LAST_LOAD_META.get("multi_bin_materials", 0),
        "decoded_bin_slots_total": sum(len(v) for v in decoded_bins.values()),
        "materials_with_line": sum(1 for m in materials if m["material_id"] in material_to_line),
        "mrp_controllers_total": len(mrp_to_line),
        "warehouse_positions": len(wh.positions),
        "warehouse_pdf_capacity": wh.pallet_capacity_from_pdf,
    }

    if write_stats:
        os.makedirs("output", exist_ok=True)
        with open("output/preprocess_stats.json", "w") as f:
            json.dump(stats, f, indent=2, ensure_ascii=False)

    return {
        "materials": materials,
        "consumption": consumption,
        "sap_master": sap_master,
        "storage_bins": storage_bins,
        "decoded_bins": decoded_bins,
        "kardex_materials": kardex_materials,
        "mrp_to_line": mrp_to_line,
        "material_to_line": material_to_line,
        "stats": stats,
    }


# Backwards-compat shim for any caller that still imports load_all_data.
def load_all_data(filepath=None):
    return {
        "abc_analizi": load_abc_analizi(filepath),
        "consumption": load_consumption(filepath),
        "sap_master": load_sap_master(filepath),
    }
