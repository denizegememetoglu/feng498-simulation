"""F400 Kit Cansu Nehir time-motion study extractor.

The workbook is a video-annotated cycle log from the F400 kitting station.
Each row is one micro-activity observed in DJI footage with elapsed time,
description, and NVA classification (CNVA/FNVA/Diğer NVA/DT NVA).

We extract per-activity duration distributions and map them to the
config.py timing constants. F400 is the largest line (~48k goods-issues
in ZWM92) so its measurements are a defensible proxy for line-agnostic
operator timing — note in ASSUMPTIONS.md §18 that other lines are
extrapolated from F400 until a per-line study is done.
"""

import json
import os
import re
import statistics
from datetime import time
from pathlib import Path
from typing import Iterable

import openpyxl


_F400_CANDIDATES = [
    Path(__file__).resolve().parent.parent / "data" / "F400 Kit Cansu Nehir.xlsx",
    Path.home() / "Downloads" / "F400 Kit Cansu Nehir.xlsx",
]
F400_KIT_FILE = os.environ.get("F400_KIT_FILE") or next(
    (str(p) for p in _F400_CANDIDATES if p.exists()), str(_F400_CANDIDATES[0])
)
SHEETS = ["Zaman Analizi Formatı", "Zaman Analizi Formatı (2)", "F400 Kit", "ist 1.1.2"]


def _time_to_seconds(t) -> float | None:
    if t is None:
        return None
    if isinstance(t, time):
        return t.hour * 3600 + t.minute * 60 + t.second + t.microsecond / 1e6
    if isinstance(t, (int, float)):
        if 0 <= t < 1:
            return float(t) * 86400.0
        if 1 <= t < 86400:
            return float(t)
    return None


def _find_desc_column(ws) -> int:
    """The description header 'Açıklama' is at row 7 or 8 depending on sheet.
    Returns column index (0-based)."""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for i, v in enumerate(row):
            if isinstance(v, str) and v.strip().lower() == "açıklama":
                return i
    return -1


def _find_elapsed_column(ws, desc_col: int) -> int:
    """Elapsed is the time(...) cell immediately left of description (col-2 usually).
    Returns column index of elapsed time."""
    # The elapsed column is the second time() column to the left of description.
    # Layout 1: cumul(col2), _, elapsed(col4), _, desc(col6)
    # Layout 2 (F400 Kit): cumul(col3), _, elapsed(col5), _, desc(col7)
    return desc_col - 2


CATEGORY_PATTERNS = [
    # Order matters: more specific patterns must come before general ones.
    ("rt_travel",          re.compile(r"(reach\s*truck.*(hareket|yürüyüş|bin|taşı|aran)|"
                                       r"boş\s*reach\s*truck)", re.IGNORECASE)),
    ("rt_pick",            re.compile(r"(reach\s*truckla|reach\s*trucktan|reach\s*truck'tan|"
                                       r"rt.*malzeme.*al)", re.IGNORECASE)),
    ("descend_rt",         re.compile(r"(inilmesi|trucktan\s*in)", re.IGNORECASE)),
    ("ascend_rt",          re.compile(r"(çıkılm|binilm)", re.IGNORECASE)),
    ("rf_scan",            re.compile(r"(rf\s*terminal|barkod|irsaliye)", re.IGNORECASE)),
    ("place_on_cart",      re.compile(r"(arabaya.*bırak|kenara.*bırak|yerleştir|"
                                       r"yer\s*arabasına|rafa\s*yeri\s*bırak)", re.IGNORECASE)),
    ("cart_move",          re.compile(r"(kit\s*arab|araba.*hareket|araba.*çek|"
                                       r"araba.*konul|araba.*düzen|araba.*taşın|"
                                       r"araba\s*tutacağ)", re.IGNORECASE)),
    # manual_pick = at-bin handling: physically taking/handling material from the
    # rack, cutting strap/wrap, etc. "Elleçleme" (handling) is the dominant
    # at-bin verb and was previously unclassified.
    ("manual_pick",        re.compile(r"(malzeme.*al(ım|ın|m|ar)|palet.*al|"
                                       r"raftan.*elleçlen|malzeme.*elleçlen|"
                                       r"şerid.*kes|streç.*kes|"
                                       r"strec.*kes|streç.*at)", re.IGNORECASE)),
    ("walk_corridor",      re.compile(r"(yürüyüş|yürüme)", re.IGNORECASE)),
    ("talking",            re.compile(r"konuşma", re.IGNORECASE)),
    ("wait",               re.compile(r"(bekleme|bekliyor|durma)", re.IGNORECASE)),
    ("check_quality",      re.compile(r"(kontrol|kalit|hata)", re.IGNORECASE)),
]


def _classify(desc: str) -> str | None:
    if not desc:
        return None
    for name, rx in CATEGORY_PATTERNS:
        if rx.search(desc):
            return name
    return None


def _iter_observations(ws) -> Iterable[tuple[float, str, tuple[float, float, float, float]]]:
    """Yield (elapsed_seconds, description, (cnva, fnva, other_nva, dt_nva))
    per data row. NVA fields are float minutes (sum of categories).
    """
    desc_col = _find_desc_column(ws)
    if desc_col < 0:
        return
    elapsed_col = _find_elapsed_column(ws, desc_col)
    # The NVA columns CNVA/FNVA/Diğer/DT live at desc_col+1..desc_col+4
    cnva_col, fnva_col, other_col, dt_col = (desc_col + 2, desc_col + 3,
                                              desc_col + 4, desc_col + 5)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if elapsed_col >= len(row) or desc_col >= len(row):
            continue
        elapsed = _time_to_seconds(row[elapsed_col])
        desc = row[desc_col]
        if elapsed is None or elapsed <= 0:
            continue
        if not isinstance(desc, str) or not desc.strip():
            continue
        nva_vals = []
        for c in (cnva_col, fnva_col, other_col, dt_col):
            v = row[c] if 0 <= c < len(row) else None
            try:
                nva_vals.append(float(v) if v else 0.0)
            except (TypeError, ValueError):
                nva_vals.append(0.0)
        yield elapsed, desc.strip(), tuple(nva_vals)


def extract(path: str = F400_KIT_FILE) -> dict:
    """Walk all time-study sheets, classify rows, return aggregate stats."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    by_cat: dict[str, list[float]] = {n: [] for n, _ in CATEGORY_PATTERNS}
    unclassified: list[tuple[str, float]] = []
    total_observed = 0
    nva_totals = {"cnva": 0.0, "fnva": 0.0, "other_nva": 0.0, "dt_nva": 0.0,
                  "all_observed_min": 0.0}

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for elapsed_s, desc, (c, f, o, d) in _iter_observations(ws):
            total_observed += 1
            elapsed_min = elapsed_s / 60.0
            nva_totals["cnva"] += c
            nva_totals["fnva"] += f
            nva_totals["other_nva"] += o
            nva_totals["dt_nva"] += d
            nva_totals["all_observed_min"] += elapsed_min
            cat = _classify(desc)
            if cat:
                by_cat[cat].append(elapsed_s)
            else:
                unclassified.append((desc, elapsed_s))
    wb.close()

    def _stats(samples: list[float]) -> dict:
        if not samples:
            return {"n": 0, "mean_s": None, "median_s": None,
                    "p95_s": None, "stdev_s": None,
                    "mean_min": None, "median_min": None}
        srt = sorted(samples)
        n = len(samples)
        mean_s = statistics.fmean(samples)
        med = statistics.median(samples)
        p95 = srt[min(n - 1, int(0.95 * (n - 1)))]
        sd = statistics.pstdev(samples) if n > 1 else 0.0
        return {
            "n": n,
            "mean_s": round(mean_s, 3),
            "median_s": round(med, 3),
            "p95_s": round(p95, 3),
            "stdev_s": round(sd, 3),
            "mean_min": round(mean_s / 60.0, 4),
            "median_min": round(med / 60.0, 4),
        }

    cat_stats = {cat: _stats(samples) for cat, samples in by_cat.items()}

    total_min = nva_totals["all_observed_min"] or 1.0
    nva_breakdown_pct = {
        "cnva_pct": round(100.0 * nva_totals["cnva"] / total_min, 2),
        "fnva_pct": round(100.0 * nva_totals["fnva"] / total_min, 2),
        "other_nva_pct": round(100.0 * nva_totals["other_nva"] / total_min, 2),
        "dt_nva_pct": round(100.0 * nva_totals["dt_nva"] / total_min, 2),
    }

    # Map to config.py constants — see ASSUMPTIONS.md §18.
    # OPERATOR_PICK_TIME = mean(manual_pick + rf_scan)
    operator_pick_s = _safe_mean([cat_stats["manual_pick"]["mean_s"],
                                  cat_stats["rf_scan"]["mean_s"]])
    operator_pick_min = (operator_pick_s / 60.0) if operator_pick_s else None

    # MANUAL_PICK_TIME_PENALTY = extra cost for non-corridor manual pick.
    # Approximate as walk_corridor mean — the walking detour without RT assist.
    # Conservative UPPER-bound proxy: full corridor traversal overstates the
    # marginal penalty over baseline pick by ~30%. See ASSUMPTIONS §19.
    penalty_s = _safe_mean([cat_stats["walk_corridor"]["mean_s"]])
    penalty_min = (penalty_s / 60.0) if penalty_s else None

    # REACH_TRUCK_PICK_PLACE_TIME = mean rt_pick
    rt_pick_s = cat_stats["rt_pick"]["mean_s"]
    rt_pick_min = (rt_pick_s / 60.0) if rt_pick_s else None

    # KARDEX_PICK_TIME — kept book value (Kardex isn't in F400 footage)
    # KARDEX_CAROUSEL_TIME — kept book value

    config_mapping = {
        "OPERATOR_PICK_TIME": round(operator_pick_min, 4) if operator_pick_min else None,
        "MANUAL_PICK_TIME_PENALTY": round(penalty_min, 4) if penalty_min else None,
        "REACH_TRUCK_PICK_PLACE_TIME": round(rt_pick_min, 4) if rt_pick_min else None,
    }

    return {
        "source_file": path,
        "sheets_processed": SHEETS,
        "total_observations": total_observed,
        "total_observed_minutes": round(total_min, 2),
        "nva_breakdown_pct": nva_breakdown_pct,
        "category_stats": cat_stats,
        "config_mapping": config_mapping,
        "unclassified_sample": unclassified[:30],
        "unclassified_count": len(unclassified),
    }


def _safe_mean(vals: list[float | None]) -> float | None:
    filt = [v for v in vals if v is not None]
    if not filt:
        return None
    return statistics.fmean(filt)


def save(out_path: str = "output/timing_study_f400.json", path: str = F400_KIT_FILE):
    result = extract(path)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(result, f, indent=2, ensure_ascii=False, default=str)
    return result


if __name__ == "__main__":
    r = save()
    print(json.dumps({
        "total_obs": r["total_observations"],
        "total_min": r["total_observed_minutes"],
        "nva_pct": r["nva_breakdown_pct"],
        "config_mapping": r["config_mapping"],
        "category_counts": {k: v["n"] for k, v in r["category_stats"].items()},
    }, indent=2, ensure_ascii=False))
